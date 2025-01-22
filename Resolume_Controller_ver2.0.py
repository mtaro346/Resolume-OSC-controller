import os
import time
from datetime import datetime, timedelta
from pythonosc import udp_client
import csv
from pathlib import Path
import logging
import sys
from openpyxl import Workbook
from openpyxl.styles import numbers

class ResolumeController:
    def __init__(self, target_time_file='target_time_template.csv', osc_ip='127.0.0.1', osc_port=7000):
        # ログディレクトリの作成
        os.makedirs('logs', exist_ok=True)
        
        # ログハンドラーの設定
        file_handler = logging.FileHandler(filename='logs/resolume_controller.log', encoding='utf-8')
        file_handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
        
        console_handler = logging.StreamHandler()
        console_handler.setFormatter(logging.Formatter('%(message)s'))
        
        # ルートロガーの設定
        logging.root.setLevel(logging.INFO)
        logging.root.addHandler(file_handler)
        logging.root.addHandler(console_handler)
        
        try:
            # CSVファイルの検索（任意の名前のCSVファイルを検索）
            csv_files = list(Path('.').glob('*.csv'))
            excel_exists = os.path.exists('target_time_template.xlsx')
            
            if len(csv_files) > 1:
                # 複数のCSVファイルが存在する場合はエラー
                error_msg = (
                    f"\n複数のCSVファイルが検出されました:\n"
                    f"{', '.join([f.name for f in csv_files])}\n"
                    f"\n処理を続けるにはどちらかのファイルを削除するか\n"
                    f"別のフォルダに移動してください。\n"
                    f"\nプログラムを終了します。"
                )
                print("\n" + "="*50)
                logging.error(error_msg)
                print("="*50 + "\n")
                input("Enterキーを押して終了してください...")
                sys.exit(1)
            elif len(csv_files) == 1:
                # 見つかったCSVファイルを使用（名前は問わない）
                self.target_time_file = str(csv_files[0])
                logging.info(f"CSVファイルを使用します: {self.target_time_file}")
                
                # CSVはあるがExcelがない場合はテンプレート作成
                if not excel_exists:
                    if self.create_template():
                        logging.info("テンプレートファイルを作成しました")
            else:
                # CSVファイルが見つからない場合
                if excel_exists:
                    # Excelのみある場合
                    logging.warning("CSVファイルが見つかりません")
                    logging.info("エクセルファイルからCSVファイルを作成してください")
                    input("Enterキーを押して終了してください...")
                    sys.exit(1)
                else:
                    # 両方ない場合は両方作成
                    self.target_time_file = target_time_file
                    self.create_template()
                    self.create_default_csv()
                    logging.info("テンプレートファイルとCSVファイルを作成しました")

            self.osc_client = udp_client.SimpleUDPClient(osc_ip, osc_port)
            self.play_times = self.get_play_times()
            
        except Exception as e:
            error_msg = f"初期化中にエラーが発生しました: {str(e)}"
            logging.error(error_msg)
            print("\n" + "="*50)
            print(error_msg)
            print("="*50 + "\n")
            input("Enterキーを押して終了してください...")
            sys.exit(1)

    def create_template(self):
        template_file = 'target_time_template.xlsx'
        if not os.path.exists(template_file):
            wb = Workbook()
            ws = wb.active
            
            # ヘッダーを設定
            headers = ['column', 'yyyy/mm/dd', 'hh:mm:ss']
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # サンプルデータを設定
            sample_data = [
                [1, '2025/01/22', '16:45:00'],
                [2, '2025/01/23', '14:30:00']
            ]
            
            for row_idx, row_data in enumerate(sample_data, 2):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    if col_idx == 2:  # 日付列
                        cell.number_format = 'yyyy/mm/dd'
                    elif col_idx == 3:  # 時刻列
                        cell.number_format = 'hh:mm:ss'
            
            # 列幅の調整
            for column in ws.columns:
                max_length = 0
                column = list(column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column[0].column_letter].width = adjusted_width
            
            wb.save(template_file)
            return True
        return False

    def create_default_csv(self):
        # CSVファイル作成を別メソッドに分離
        with open(self.target_time_file, 'w', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(['column', 'yyyy/mm/dd', 'hh:mm:ss'])
            writer.writerow([1, '2025/01/22', '16:45:00'])
            writer.writerow([2, '2025/01/23', '14:30:00'])
        logging.info(f"デフォルトのCSVファイルを作成しました: {self.target_time_file}")

    def validate_date(self, date_str):
        try:
            return datetime.strptime(date_str, '%Y/%m/%d')
        except ValueError:
            return None

    def validate_time(self, time_str):
        try:
            hours, minutes, seconds = map(int, time_str.split(':'))  # 秒も取得
            if 0 <= hours <= 23 and 0 <= minutes <= 59 and 0 <= seconds <= 59:  # 秒の範囲もチェック
                return True
            return False
        except:
            return False

    def get_play_times(self):
        play_times = []
        with open(self.target_time_file, 'r') as f:
            reader = csv.DictReader(f)
            for row_num, row in enumerate(reader, start=2):
                try:
                    # データの検証
                    if not row['column'].isdigit():
                        raise ValueError(f"無効な列番号です: {row['column']}")
                    
                    column = int(row['column'])
                    date_obj = self.validate_date(row['yyyy/mm/dd'])
                    if not date_obj:
                        raise ValueError(f"無効な日付形式です: {row['yyyy/mm/dd']}")
                    
                    if not self.validate_time(row['hh:mm:ss']):
                        raise ValueError(f"無効な時刻形式です: {row['hh:mm:ss']}")
                    
                    hours, minutes, seconds = map(int, row['hh:mm:ss'].split(':'))  # 秒も取得
                    play_time = date_obj.replace(hour=hours, minute=minutes, second=seconds, microsecond=0)
                    
                    if play_time < datetime.now():
                        logging.warning(f"行 {row_num}: 過去の日時のためスキップしました: {play_time}")
                        continue
                    
                    play_times.append((play_time, column))
                    
                except Exception as e:
                    logging.error(f"行 {row_num}: {str(e)}")
                    continue

        # 時間順にソート
        play_times.sort(key=lambda x: x[0])
        return play_times

    def send_osc_command(self, column):
        # OSCコマンドを送信するメソッド。指定されたコラムを選択するコマンドを送信します。
        try:
            self.osc_client.send_message(f'/composition/columns/{column}/connect', 1)
            time.sleep(0.1)  # コマンドの処理が間に合うように短い遅延を挿入

            print(f'\nコラム{column}にコマンドを送信しました。')
            return True
        except Exception as e:
            print(f'エラーが発生しました: {e}')
            return False

    def run(self):
        # メインの実行メソッド。指定された再生時間まで待機し、OSCコマンドを送信します。
        try:
            for i, (play_time, column) in enumerate(self.play_times):
                while True:
                    now = datetime.now()
                    remaining = play_time - now
                    remaining_seconds = int(remaining.total_seconds())
                    if remaining_seconds <= 0:
                        for attempt in range(2):  # 2回まで再送信を試みる
                            if self.send_osc_command(column):
                                break
                            print(f'再試行 {attempt + 1}/2')
                            time.sleep(1)
                        else:
                            with open('error_log.txt', 'a') as log_file:
                                log_file.write(f'{datetime.now()}: コマンド送信に失敗しました。\n')
                            print('エラーが発生しました。エラーログを確認してください。')
                        
                        # 最後のスケジュールが終了したかチェック
                        if i == len(self.play_times) - 1:
                            print("\n\nすべてのスケジュールの再生が終了しました。")
                            input("Enterキーを押して終了します...")
                        break  # 次のスケジュールに進む

                    # 残り時間の計算
                    days = remaining_seconds // 86400  # 日数を計算
                    remaining_seconds %= 86400  # 日数を除いた残り秒数
                    hours, remainder = divmod(remaining_seconds, 3600)
                    minutes, seconds = divmod(remainder, 60)

                    # 表示文字列の作成
                    current_time_str = now.strftime('%Y/%m/%d %H:%M:%S')
                    play_time_str = play_time.strftime('%Y/%m/%d %H:%M:%S')
                    
                    # 残り時間の文字列（日数がある場合のみ日数を表示）
                    if days > 0:
                        remaining_str = f'{days}日と{hours:02d}時間{minutes:02d}分{seconds:02d}秒'
                    else:
                        remaining_str = f'{hours:02d}時間{minutes:02d}分{seconds:02d}秒'

                    print(f'\r現在時刻: {current_time_str} | 再生時刻: {play_time_str} | '
                          f'再生コラム: {column} | 残り時間: {remaining_str}', end='')
                    time.sleep(1)
        except KeyboardInterrupt:
            print('\nプログラムが中断されました。')

if __name__ == '__main__':
    try:
        controller = ResolumeController()
        controller.run()
    except KeyboardInterrupt:
        print('\nプログラムが中断されました。')
        input("Enterキーを押して終了してください...")
        sys.exit(0)
    except Exception as e:
        logging.error(f"予期せぬエラーが発生しました: {str(e)}")
        print("\n" + "="*50)
        print(f"予期せぬエラーが発生しました: {str(e)}")
        print("詳細はログファイルを確認してください。")
        print("="*50 + "\n")
        input("Enterキーを押して終了してください...")
        sys.exit(1) 